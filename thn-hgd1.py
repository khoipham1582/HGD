from selenium import  webdriver
from time import sleep
import random
from selenium.webdriver.common.keys import Keys
import pandas as pd
browser = webdriver.Chrome(executable_path="chromedriver.exe")
from selenium.webdriver.common.by import By
from selenium.common.exceptions import NoSuchElementException    
#from xlwt import Workbook
from openpyxl import Workbook
from datetime import datetime
#xl=pd.ExcelFile("ds2.xlsx")
#df=pd.read_excel(xl,0,header=None)
now=datetime.now()
#print(now.strftime("%d-%m-%H-%M-%S"))
k=2
j=99999300
wb = Workbook()
ws=wb.active
#ws=wb.create_sheet(title="Ketqua")
ws['A1']='STT'
ws['B1']='Mã số BHXH'
ws['C1']='Họ tên'
ws['D1']='Ngày sinh'
ws['E1']='Giới tính'
ws['F1']='Mã tỉnh KS'
ws['G1']='Mã huyện KS'
ws['H1']='Mã xã KS'
ws['I1']='Mã tỉnh HK'
ws['J1']='Mã huyện HK'
ws['K1']='Mã xã HK'
ws['L1']='QH chủ hộ'
ws['M1']='Loại phát sinh'
ws['N1']='Trạng thái'
ws['O1']='Nơi khai sinh'
ws['P1']='Số CMTND'
ws['Q1']='Ghi chú'
ws['R1']='Mã hộ'
'''
#ws = wb.add_sheet('ketqua') 
ws.append(0, 0, 'STT')
ws.append(0, 1, 'Mã số BHXH')
ws.append(0, 2, 'Họ tên')
ws.append(0, 3, 'Ngày sinh')
ws.append(0, 4, 'Giới tính')
ws.append(0, 5, 'Mã tỉnh KS')
ws.append(0, 6, 'Mã huyện KS')
ws.append(0, 7, 'Mã xã KS')
ws.append(0, 8, 'Mã tỉnh HK')
ws.append(0, 9, 'Mã huyện HK')
ws.append(0, 10, 'Mã xã HK')
ws.append(0, 11, 'QH chủ hộ')
ws.append(0, 12, 'Loại phát sinh')
ws.append(0, 13, 'Trạng thái')
ws.append(0, 14, 'Nơi khai sinh')
ws.append(0, 15, 'Số CMTND')
ws.append(0, 16, 'Ghi chú')
ws.append(0, 17, 'Mã hộ')
ws.append(0, 18, 'Số thành viên')
'''
#soluong=len(df.iloc[:,0]
# đăng nhập hgđ
browser.get("https://tst.bhxh.gov.vn/")
#sleep(2)
#browser.find_element_by_id("bhxhMa").send_keys("06401")
browser.find_element_by_id("username").send_keys("cntt@gialai.vss.gov.vn")
browser.find_element_by_id("password").send_keys("H0G1@D1nh")
browser.find_element_by_id("password").send_keys(Keys.ENTER)
#browser.find_element_by_id("password").send_keys(Keys.F5)
#print("Bạn có 40s để đăng nhập.. !")
#sleep(40)
browser.get("https://tst.baohiemxahoi.gov.vn/#/qlNhanKhau")
#browser.find_element_by_xpath("/html/body/div[3]/div[1]/div/form/div[2]/div[2]/input").get_attribute('checked')
#tctq=browser.find_elements_by_css_selector("body > div.container > div.ng-scope > div > form > div.row.padding-10-top > div.col-sm-6.text-right > input")
#tctq=browser.find_elements_by_class_name("ng-pristine ng-valid ng-touched")
#tctq[1].click()
#tctq[1].click()
#print(*tctq,sep=",")
#sleep(3)
#xl=pd.ExcelFile("ds2.xlsx")
#df=pd.read_excel(xl,0,header=None)
now= now.strftime("%d-%m-%H-%M-%S")  
#print(df.head())
#for i in range(len(df.iloc[:,0])):
def check_exists_by_xpath(xpath):
    try:
        browser.find_element_by_xpath(xpath)
    except NoSuchElementException:
        return False
    return True
def traHGD(maho): 
    try:   
        sleep(1)
        global k
        mh=browser.find_element_by_xpath("/html/body/div[3]/div[1]/div[2]/div[2]/div/fieldset/div[3]/div/div[1]/div/div/input")
        mh.send_keys(maho)
        #browser.execute_script("arguments[0].value="+maho,mh)
        #print("Tra HGD :"+maho)
        browser.find_element_by_xpath("/html/body/div[3]/div[1]/div[2]/div[2]/div/div/div/button[2]/span[2]").click()
        sleep(1)
        table=browser.find_element_by_xpath("/html/body/div[3]/div[1]/div[2]/div[3]/div/div/table")
        tr=table.find_elements_by_tag_name("tr")
        if(len(tr)>1):
            for i in range(len(tr)-1):
                #print(i+1)
                maso=tr[i+1].find_elements_by_tag_name("td")[3].text
                ht=tr[i+1].find_elements_by_tag_name("td")[4].text
                ns=tr[i+1].find_elements_by_tag_name("td")[5].text
                gt=tr[i+1].find_elements_by_tag_name("td")[6].text
                tks=tr[i+1].find_elements_by_tag_name("td")[7].text
                hks=tr[i+1].find_elements_by_tag_name("td")[8].text
                xks=tr[i+1].find_elements_by_tag_name("td")[9].text
                thk=tr[i+1].find_elements_by_tag_name("td")[10].text
                hhk=tr[i+1].find_elements_by_tag_name("td")[11].text
                xhk=tr[i+1].find_elements_by_tag_name("td")[12].text
                qh=tr[i+1].find_elements_by_tag_name("td")[13].text
                lps=tr[i+1].find_elements_by_tag_name("td")[14].text
                tt=tr[i+1].find_elements_by_tag_name("td")[15].text
                nks=tr[i+1].find_elements_by_tag_name("td")[16].text
                cmnd=tr[i+1].find_elements_by_tag_name("td")[17].text
                ghichu=tr[i+1].find_elements_by_tag_name("td")[18].text
                ws['A'+str(k)]=i+1
                ws['B'+str(k)]=maso
                ws['C'+str(k)]=ht
                ws['D'+str(k)]=ns
                ws['E'+str(k)]=gt
                ws['F'+str(k)]=tks
                ws['G'+str(k)]=hks
                ws['H'+str(k)]=xks
                ws['I'+str(k)]=thk
                ws['J'+str(k)]=hhk
                ws['K'+str(k)]=xhk
                ws['L'+str(k)]=qh
                ws['M'+str(k)]=lps
                ws['N'+str(k)]=tt
                ws['O'+str(k)]=nks
                ws['P'+str(k)]=cmnd
                ws['Q'+str(k)]=ghichu
                ws['R'+str(k)]=maho
                ws['S'+str(k)]=len(tr)-1
                k=k+1
                '''
                ws.append(k, 0, i+1)
                ws.append(k, 1, maso)
                ws.append(k, 2, ht)
                ws.append(k, 3, ns)
                ws.append(k, 4, gt)
                ws.append(k, 5, tks)
                ws.append(k, 6, hks)
                ws.append(k, 7, xks)
                ws.append(k, 8, thk)
                ws.append(k, 9, hhk)
                ws.append(k, 10, xhk)
                ws.append(k, 11, qh)
                ws.append(k, 12, lps)
                ws.append(k, 13, tt)
                ws.append(k, 14, nks)
                ws.append(k, 15, cmnd)
                ws.append(k, 16, ghichu)
                ws.append(k, 17, maho)
                '''
                
        else:
            ws['Q'+str(k)]='Không có nhân khẩu'
            ws['R'+str(k)]=maho
            #ws.append(k, 17, maho)
            #ws.append(k, 16, "Không có nhân khẩu")
            k=k+1
        browser.execute_script("arguments[0].value=''",mh)
    except:
        wb.save(filename ='HGD'+now+'.xlsx')
    #browser.find_element_by_xpath("/html/body/div[3]/div[1]/div[2]/div[2]/div/fieldset/div[3]/div/div[1]/div/div/input").send_keys("") 
while(j>97000000 or k>999999):
    #trahgd(i,df.at[i,0],df.at[i,1])
    #sleep(1)
    traHGD('01'+str(j))
    j=j-1
    
    #print('01'+i)
wb.save(filename ='HGD'+now+'.xlsx')
browser.close()